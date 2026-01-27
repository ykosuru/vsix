#!/usr/bin/env python3
"""
PDF/XSD/Excel to Markdown Converter

Usage:
    python pdf2md.py input.pdf                    # Output: input.md + input_images/
    python pdf2md.py input.xsd                    # Output: input.md (formatted XSD)
    python pdf2md.py input.xlsx                   # Output: input.md (tables)
    python pdf2md.py input.pdf -o output.md       # Custom output name
    python pdf2md.py input.pdf --no-images        # Skip image extraction
    python pdf2md.py *.pdf *.xsd *.xlsx           # Batch convert
    python pdf2md.py docs/ -r                     # Recursive directory

Requirements:
    pip install pymupdf pillow openpyxl xlrd
"""

import argparse
import os
import re
import sys
import xml.etree.ElementTree as ET
from pathlib import Path
from typing import Optional

try:
    import fitz  # pymupdf
    HAS_PYMUPDF = True
except ImportError:
    HAS_PYMUPDF = False
    fitz = None  # type: ignore

try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

try:
    import xlrd
    HAS_XLRD = True
except ImportError:
    HAS_XLRD = False


def extract_images(page, page_num: int, output_dir: Path) -> list[dict]:
    """Extract images from a PDF page."""
    images = []
    image_list = page.get_images(full=True)
    
    for img_index, img_info in enumerate(image_list):
        xref = img_info[0]
        
        try:
            base_image = page.parent.extract_image(xref)
            if not base_image:
                continue
                
            image_bytes = base_image["image"]
            image_ext = base_image["ext"]
            
            # Generate filename
            img_filename = f"page{page_num + 1}_img{img_index + 1}.{image_ext}"
            img_path = output_dir / img_filename
            
            # Save image
            with open(img_path, "wb") as f:
                f.write(image_bytes)
            
            images.append({
                "filename": img_filename,
                "path": str(img_path),
                "page": page_num + 1,
                "index": img_index + 1
            })
            
        except Exception as e:
            print(f"  Warning: Could not extract image {img_index + 1} on page {page_num + 1}: {e}", file=sys.stderr)
    
    return images


def clean_text(text: str) -> str:
    """Clean extracted text for markdown."""
    # Remove excessive whitespace but preserve paragraph breaks
    text = re.sub(r'\n{3,}', '\n\n', text)
    
    # Fix common PDF extraction issues
    text = re.sub(r'(?<=[a-z])-\n(?=[a-z])', '', text)  # Fix hyphenation
    text = re.sub(r'(?<=[.!?])\n(?=[A-Z])', '\n\n', text)  # Add paragraph breaks after sentences
    
    return text.strip()


def detect_headings(text: str) -> str:
    """Try to detect and format headings."""
    lines = text.split('\n')
    result = []
    
    for i, line in enumerate(lines):
        stripped = line.strip()
        
        # Skip empty lines
        if not stripped:
            result.append(line)
            continue
        
        # Detect potential headings (short, possibly uppercase, no ending punctuation)
        is_short = len(stripped) < 80
        is_uppercase = stripped.isupper() and len(stripped) > 3
        no_end_punct = not stripped.endswith(('.', ',', ';', ':'))
        next_is_empty = i + 1 < len(lines) and not lines[i + 1].strip()
        prev_is_empty = i > 0 and not lines[i - 1].strip()
        
        if is_short and no_end_punct and (prev_is_empty or i == 0):
            if is_uppercase:
                # Major heading
                result.append(f"\n## {stripped.title()}\n")
            elif stripped[0].isupper() and len(stripped) < 50 and next_is_empty:
                # Minor heading
                result.append(f"\n### {stripped}\n")
            else:
                result.append(line)
        else:
            result.append(line)
    
    return '\n'.join(result)


def convert_xsd_to_markdown(
    xsd_path: Path,
    output_path: Optional[Path] = None
) -> dict:
    """
    Convert an XSD file to Markdown with structured documentation.
    
    Returns:
        dict with keys: output_path, element_count, type_count
    """
    xsd_path = Path(xsd_path)
    
    if not xsd_path.exists():
        raise FileNotFoundError(f"XSD not found: {xsd_path}")
    
    # Determine output path
    if output_path is None:
        output_path = xsd_path.with_suffix('.md')
    else:
        output_path = Path(output_path)
    
    print(f"Converting: {xsd_path}")
    
    # Read XSD content
    with open(xsd_path, 'r', encoding='utf-8') as f:
        xsd_content = f.read()
    
    # Parse XSD
    try:
        root = ET.fromstring(xsd_content)
    except ET.ParseError as e:
        # If parsing fails, just wrap raw content
        print(f"  Warning: Could not parse XSD structure: {e}")
        markdown = generate_raw_xsd_markdown(xsd_path, xsd_content)
        with open(output_path, 'w', encoding='utf-8') as f:
            f.write(markdown)
        return {"output_path": str(output_path), "element_count": 0, "type_count": 0}
    
    # Extract namespace
    ns = {'xs': 'http://www.w3.org/2001/XMLSchema'}
    
    # Build markdown
    markdown_content = []
    
    # Title
    title = xsd_path.stem.replace('_', ' ').replace('-', ' ').title()
    markdown_content.append(f"# {title}\n")
    markdown_content.append(f"*Schema: {xsd_path.name}*\n")
    markdown_content.append("---\n")
    
    # Extract elements
    elements = root.findall('.//xs:element', ns)
    complex_types = root.findall('.//xs:complexType', ns)
    simple_types = root.findall('.//xs:simpleType', ns)
    
    # Summary
    markdown_content.append("## Summary\n")
    markdown_content.append(f"- **Elements:** {len(elements)}")
    markdown_content.append(f"- **Complex Types:** {len(complex_types)}")
    markdown_content.append(f"- **Simple Types:** {len(simple_types)}\n")
    
    # Document elements
    if elements:
        markdown_content.append("## Elements\n")
        markdown_content.append("| Element | Type | Min | Max | Description |")
        markdown_content.append("|---------|------|-----|-----|-------------|")
        
        for elem in elements:
            name = elem.get('name', '(anonymous)')
            elem_type = elem.get('type', '(complex)')
            min_occurs = elem.get('minOccurs', '1')
            max_occurs = elem.get('maxOccurs', '1')
            
            # Try to get documentation
            doc = elem.find('.//xs:documentation', ns)
            description = doc.text.strip()[:50] if doc is not None and doc.text else ''
            
            markdown_content.append(f"| `{name}` | `{elem_type}` | {min_occurs} | {max_occurs} | {description} |")
        
        markdown_content.append("")
    
    # Document complex types
    if complex_types:
        markdown_content.append("## Complex Types\n")
        
        for ct in complex_types:
            name = ct.get('name', '(anonymous)')
            markdown_content.append(f"### {name}\n")
            
            # Get child elements
            child_elements = ct.findall('.//xs:element', ns)
            if child_elements:
                markdown_content.append("| Field | Type | Required |")
                markdown_content.append("|-------|------|----------|")
                
                for child in child_elements:
                    child_name = child.get('name', child.get('ref', '?'))
                    child_type = child.get('type', '(complex)')
                    required = 'Yes' if child.get('minOccurs', '1') != '0' else 'No'
                    markdown_content.append(f"| `{child_name}` | `{child_type}` | {required} |")
                
                markdown_content.append("")
    
    # Document simple types (enums, restrictions)
    if simple_types:
        markdown_content.append("## Simple Types\n")
        
        for st in simple_types:
            name = st.get('name', '(anonymous)')
            markdown_content.append(f"### {name}\n")
            
            # Check for enumeration
            enums = st.findall('.//xs:enumeration', ns)
            if enums:
                markdown_content.append("**Values:**")
                for enum in enums:
                    val = enum.get('value', '')
                    markdown_content.append(f"- `{val}`")
                markdown_content.append("")
            
            # Check for restrictions
            restriction = st.find('.//xs:restriction', ns)
            if restriction is not None:
                base = restriction.get('base', '')
                markdown_content.append(f"**Base type:** `{base}`\n")
                
                # Pattern
                pattern = restriction.find('xs:pattern', ns)
                if pattern is not None:
                    markdown_content.append(f"**Pattern:** `{pattern.get('value', '')}`\n")
                
                # Length restrictions
                min_len = restriction.find('xs:minLength', ns)
                max_len = restriction.find('xs:maxLength', ns)
                if min_len is not None or max_len is not None:
                    min_v = min_len.get('value', '0') if min_len is not None else '0'
                    max_v = max_len.get('value', '∞') if max_len is not None else '∞'
                    markdown_content.append(f"**Length:** {min_v} - {max_v}\n")
    
    # Include raw XSD
    markdown_content.append("## Raw Schema\n")
    markdown_content.append("```xml")
    markdown_content.append(xsd_content)
    markdown_content.append("```\n")
    
    # Write output
    final_content = '\n'.join(markdown_content)
    with open(output_path, 'w', encoding='utf-8') as f:
        f.write(final_content)
    
    print(f"  Output: {output_path}")
    print(f"  Elements: {len(elements)}, Types: {len(complex_types) + len(simple_types)}")
    
    return {
        "output_path": str(output_path),
        "element_count": len(elements),
        "type_count": len(complex_types) + len(simple_types)
    }


def generate_raw_xsd_markdown(xsd_path: Path, content: str) -> str:
    """Generate markdown with raw XSD content when parsing fails."""
    title = xsd_path.stem.replace('_', ' ').replace('-', ' ').title()
    return f"""# {title}

*Schema: {xsd_path.name}*

---

## Schema Definition

```xml
{content}
```
"""


def convert_xlsx_to_markdown(
    xlsx_path: Path,
    output_path: Optional[Path] = None
) -> dict:
    """
    Convert an XLSX file to Markdown tables.
    
    Returns:
        dict with keys: output_path, sheet_count, row_count
    """
    xlsx_path = Path(xlsx_path)
    
    if not xlsx_path.exists():
        raise FileNotFoundError(f"Excel file not found: {xlsx_path}")
    
    if output_path is None:
        output_path = xlsx_path.with_suffix('.md')
    else:
        output_path = Path(output_path)
    
    print(f"Converting: {xlsx_path}")
    
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    
    markdown_content = []
    total_rows = 0
    
    # Title
    title = xlsx_path.stem.replace('_', ' ').replace('-', ' ').title()
    markdown_content.append(f"# {title}\n")
    markdown_content.append(f"*Source: {xlsx_path.name}*\n")
    markdown_content.append("---\n")
    
    # Process each sheet
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        
        # Get data as list of rows
        rows = list(ws.iter_rows(values_only=True))
        
        # Skip empty sheets
        if not rows or all(all(cell is None for cell in row) for row in rows):
            continue
        
        # Find actual data bounds (skip empty rows/cols)
        non_empty_rows = [row for row in rows if any(cell is not None for cell in row)]
        if not non_empty_rows:
            continue
        
        markdown_content.append(f"## {sheet_name}\n")
        
        # Use first row as header
        header = non_empty_rows[0]
        data_rows = non_empty_rows[1:]
        
        # Determine column count from header
        col_count = len([c for c in header if c is not None]) or len(header)
        
        # Format header
        header_cells = [str(cell) if cell is not None else '' for cell in header[:col_count]]
        markdown_content.append("| " + " | ".join(header_cells) + " |")
        markdown_content.append("|" + "|".join(["---"] * col_count) + "|")
        
        # Format data rows
        for row in data_rows:
            cells = [str(cell) if cell is not None else '' for cell in row[:col_count]]
            # Escape pipe characters in cell content
            cells = [c.replace('|', '\\|') for c in cells]
            markdown_content.append("| " + " | ".join(cells) + " |")
            total_rows += 1
        
        markdown_content.append("")
    
    wb.close()
    
    # Write output
    final_content = '\n'.join(markdown_content)
    with open(output_path, 'w', encoding='utf-8') as f:
        f.write(final_content)
    
    print(f"  Output: {output_path}")
    print(f"  Sheets: {len(wb.sheetnames)}, Rows: {total_rows}")
    
    return {
        "output_path": str(output_path),
        "sheet_count": len(wb.sheetnames),
        "row_count": total_rows
    }


def convert_xls_to_markdown(
    xls_path: Path,
    output_path: Optional[Path] = None
) -> dict:
    """
    Convert an XLS (legacy Excel) file to Markdown tables.
    
    Returns:
        dict with keys: output_path, sheet_count, row_count
    """
    xls_path = Path(xls_path)
    
    if not xls_path.exists():
        raise FileNotFoundError(f"Excel file not found: {xls_path}")
    
    if output_path is None:
        output_path = xls_path.with_suffix('.md')
    else:
        output_path = Path(output_path)
    
    print(f"Converting: {xls_path}")
    
    wb = xlrd.open_workbook(xls_path)
    
    markdown_content = []
    total_rows = 0
    
    # Title
    title = xls_path.stem.replace('_', ' ').replace('-', ' ').title()
    markdown_content.append(f"# {title}\n")
    markdown_content.append(f"*Source: {xls_path.name}*\n")
    markdown_content.append("---\n")
    
    # Process each sheet
    for sheet_idx in range(wb.nsheets):
        ws = wb.sheet_by_index(sheet_idx)
        
        # Skip empty sheets
        if ws.nrows == 0:
            continue
        
        markdown_content.append(f"## {ws.name}\n")
        
        # Use first row as header
        if ws.nrows > 0:
            header = [str(ws.cell_value(0, col)) for col in range(ws.ncols)]
            markdown_content.append("| " + " | ".join(header) + " |")
            markdown_content.append("|" + "|".join(["---"] * ws.ncols) + "|")
        
        # Format data rows
        for row_idx in range(1, ws.nrows):
            cells = []
            for col_idx in range(ws.ncols):
                cell_value = ws.cell_value(row_idx, col_idx)
                cell_str = str(cell_value) if cell_value is not None else ''
                # Escape pipe characters
                cell_str = cell_str.replace('|', '\\|')
                cells.append(cell_str)
            markdown_content.append("| " + " | ".join(cells) + " |")
            total_rows += 1
        
        markdown_content.append("")
    
    # Write output
    final_content = '\n'.join(markdown_content)
    with open(output_path, 'w', encoding='utf-8') as f:
        f.write(final_content)
    
    print(f"  Output: {output_path}")
    print(f"  Sheets: {wb.nsheets}, Rows: {total_rows}")
    
    return {
        "output_path": str(output_path),
        "sheet_count": wb.nsheets,
        "row_count": total_rows
    }


def convert_pdf_to_markdown(
    pdf_path: Path,
    output_path: Optional[Path] = None,
    extract_images_flag: bool = True,
    detect_headings_flag: bool = True
) -> dict:
    """
    Convert a PDF file to Markdown.
    
    Returns:
        dict with keys: output_path, images_dir, image_count, page_count
    """
    pdf_path = Path(pdf_path)
    
    if not pdf_path.exists():
        raise FileNotFoundError(f"PDF not found: {pdf_path}")
    
    # Determine output paths
    if output_path is None:
        output_path = pdf_path.with_suffix('.md')
    else:
        output_path = Path(output_path)
    
    images_dir = output_path.parent / f"{output_path.stem}_images"
    
    # Create images directory if extracting images
    if extract_images_flag:
        images_dir.mkdir(parents=True, exist_ok=True)
    
    print(f"Converting: {pdf_path}")
    
    # Open PDF
    doc = fitz.open(pdf_path)
    page_count = len(doc)
    
    markdown_content = []
    all_images = []
    
    # Add title from filename
    title = pdf_path.stem.replace('_', ' ').replace('-', ' ').title()
    markdown_content.append(f"# {title}\n")
    markdown_content.append(f"*Converted from: {pdf_path.name}*\n")
    markdown_content.append("---\n")
    
    # Process each page
    for page_num in range(page_count):
        page = doc[page_num]
        
        # Extract text
        text = page.get_text("text")
        text = clean_text(text)
        
        if detect_headings_flag:
            text = detect_headings(text)
        
        # Add page marker for long documents
        if page_count > 5:
            markdown_content.append(f"\n<!-- Page {page_num + 1} -->\n")
        
        markdown_content.append(text)
        
        # Extract images
        if extract_images_flag:
            page_images = extract_images(page, page_num, images_dir)
            
            # Insert image references
            for img in page_images:
                rel_path = f"{images_dir.name}/{img['filename']}"
                markdown_content.append(f"\n![Image from page {img['page']}]({rel_path})\n")
            
            all_images.extend(page_images)
    
    doc.close()
    
    # Write markdown file
    final_content = '\n'.join(markdown_content)
    
    # Clean up excessive newlines
    final_content = re.sub(r'\n{4,}', '\n\n\n', final_content)
    
    with open(output_path, 'w', encoding='utf-8') as f:
        f.write(final_content)
    
    # Remove empty images directory if no images
    if extract_images_flag and not all_images and images_dir.exists():
        try:
            images_dir.rmdir()
        except:
            pass
    
    print(f"  Output: {output_path}")
    if all_images:
        print(f"  Images: {len(all_images)} extracted to {images_dir}")
    
    return {
        "output_path": str(output_path),
        "images_dir": str(images_dir) if all_images else None,
        "image_count": len(all_images),
        "page_count": page_count
    }


def process_directory(
    dir_path: Path,
    recursive: bool = False,
    **kwargs
) -> list[dict]:
    """Process all PDFs, XSDs, and Excel files in a directory."""
    results = []
    
    # Find all supported files
    if recursive:
        pdf_files = list(dir_path.glob("**/*.pdf"))
        xsd_files = list(dir_path.glob("**/*.xsd"))
        xlsx_files = list(dir_path.glob("**/*.xlsx"))
        xls_files = list(dir_path.glob("**/*.xls"))
    else:
        pdf_files = list(dir_path.glob("*.pdf"))
        xsd_files = list(dir_path.glob("*.xsd"))
        xlsx_files = list(dir_path.glob("*.xlsx"))
        xls_files = list(dir_path.glob("*.xls"))
    
    # Exclude .xlsx from .xls results (glob *.xls matches *.xlsx too)
    xls_files = [f for f in xls_files if f.suffix.lower() == '.xls']
    
    all_files = pdf_files + xsd_files + xlsx_files + xls_files
    
    if not all_files:
        print(f"No PDF, XSD, or Excel files found in {dir_path}")
        return results
    
    file_counts = []
    if pdf_files: file_counts.append(f"{len(pdf_files)} PDF")
    if xsd_files: file_counts.append(f"{len(xsd_files)} XSD")
    if xlsx_files: file_counts.append(f"{len(xlsx_files)} XLSX")
    if xls_files: file_counts.append(f"{len(xls_files)} XLS")
    print(f"Found {', '.join(file_counts)} files\n")
    
    for file_path in sorted(all_files):
        try:
            suffix = file_path.suffix.lower()
            if suffix == '.pdf':
                if not HAS_PYMUPDF:
                    print(f"  Skipping {file_path.name} (pymupdf not installed)")
                    continue
                result = convert_pdf_to_markdown(file_path, **kwargs)
            elif suffix == '.xsd':
                result = convert_xsd_to_markdown(file_path)
            elif suffix == '.xlsx':
                if not HAS_OPENPYXL:
                    print(f"  Skipping {file_path.name} (openpyxl not installed)")
                    continue
                result = convert_xlsx_to_markdown(file_path)
            elif suffix == '.xls':
                if not HAS_XLRD:
                    print(f"  Skipping {file_path.name} (xlrd not installed)")
                    continue
                result = convert_xls_to_markdown(file_path)
            else:
                continue
            results.append(result)
        except Exception as e:
            print(f"  Error: {e}", file=sys.stderr)
            results.append({"error": str(e), "file": str(file_path)})
        print()
    
    return results


def main():
    parser = argparse.ArgumentParser(
        description="Convert PDF, XSD, and Excel files to Markdown",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  %(prog)s document.pdf                  Convert single PDF
  %(prog)s schema.xsd                    Convert single XSD
  %(prog)s data.xlsx                     Convert single Excel file
  %(prog)s document.pdf -o output.md     Custom output name
  %(prog)s *.pdf *.xsd *.xlsx            Convert multiple files
  %(prog)s docs/ -r                      Convert all files in directory (recursive)
  %(prog)s spec.pdf --no-images          Skip image extraction (PDF only)
        """
    )
    
    parser.add_argument(
        "input",
        nargs="+",
        help="PDF/XSD/Excel file(s) or directory to convert"
    )
    
    parser.add_argument(
        "-o", "--output",
        help="Output markdown file (only for single file input)"
    )
    
    parser.add_argument(
        "--no-images",
        action="store_true",
        help="Skip image extraction (PDF only)"
    )
    
    parser.add_argument(
        "--no-headings",
        action="store_true",
        help="Skip automatic heading detection (PDF only)"
    )
    
    parser.add_argument(
        "-r", "--recursive",
        action="store_true",
        help="Process directories recursively"
    )
    
    args = parser.parse_args()
    
    kwargs = {
        "extract_images_flag": not args.no_images,
        "detect_headings_flag": not args.no_headings
    }
    
    results = []
    
    for input_path in args.input:
        path = Path(input_path)
        
        if path.is_dir():
            results.extend(process_directory(path, recursive=args.recursive, **kwargs))
        elif path.exists():
            output = Path(args.output) if args.output and len(args.input) == 1 else None
            try:
                suffix = path.suffix.lower()
                if suffix == '.pdf':
                    if not HAS_PYMUPDF:
                        print("Error: pymupdf not installed. Run: pip install pymupdf", file=sys.stderr)
                        continue
                    result = convert_pdf_to_markdown(path, output_path=output, **kwargs)
                elif suffix == '.xsd':
                    result = convert_xsd_to_markdown(path, output_path=output)
                elif suffix == '.xlsx':
                    if not HAS_OPENPYXL:
                        print("Error: openpyxl not installed. Run: pip install openpyxl", file=sys.stderr)
                        continue
                    result = convert_xlsx_to_markdown(path, output_path=output)
                elif suffix == '.xls':
                    if not HAS_XLRD:
                        print("Error: xlrd not installed. Run: pip install xlrd", file=sys.stderr)
                        continue
                    result = convert_xls_to_markdown(path, output_path=output)
                else:
                    print(f"Unsupported file type: {suffix}", file=sys.stderr)
                    continue
                results.append(result)
            except Exception as e:
                print(f"Error: {e}", file=sys.stderr)
                results.append({"error": str(e), "file": str(path)})
        elif '*' in input_path:
            # Handle glob patterns
            import glob
            for matched_path in glob.glob(input_path):
                matched = Path(matched_path)
                suffix = matched.suffix.lower()
                if suffix in ('.pdf', '.xsd', '.xlsx', '.xls'):
                    try:
                        if suffix == '.pdf':
                            if not HAS_PYMUPDF:
                                continue
                            result = convert_pdf_to_markdown(matched, **kwargs)
                        elif suffix == '.xsd':
                            result = convert_xsd_to_markdown(matched)
                        elif suffix == '.xlsx':
                            if not HAS_OPENPYXL:
                                continue
                            result = convert_xlsx_to_markdown(matched)
                        elif suffix == '.xls':
                            if not HAS_XLRD:
                                continue
                            result = convert_xls_to_markdown(matched)
                        results.append(result)
                    except Exception as e:
                        print(f"Error: {e}", file=sys.stderr)
        else:
            print(f"Not found: {input_path}", file=sys.stderr)
    
    # Summary
    if len(results) > 1:
        success = sum(1 for r in results if "error" not in r)
        total_images = sum(r.get("image_count", 0) for r in results if "error" not in r)
        print(f"\n{'='*40}")
        print(f"Converted: {success}/{len(results)} files")
        if total_images:
            print(f"Total images extracted: {total_images}")


if __name__ == "__main__":
    main()
